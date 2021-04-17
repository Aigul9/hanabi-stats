import pandas as pd
import xlsxwriter
from openpyxl import load_workbook
from py.constants import username


class UserStat:
    def __init__(self, game_id, count, score, variant, date, players, other_scores, suits, max_score):
        self.game_id = game_id
        self.count = count
        self.score = score
        self.variant = variant
        self.date = date
        self.players = players
        self.other_scores = other_scores
        self.suits = suits
        self.max_score = max_score


def get_percentage(value, total_value):
    return round(value * 100 / total_value, 2)


def get_double_percentage(totals_list):
    value1, value2, total_value = totals_list
    return get_percentage(value1, total_value), get_percentage(value2, total_value)


def get_totals(stat_list):
    total = len(stat_list)
    total_wins = len([row for row in stat_list if row.score == row.max_score])
    total_losses = total - total_wins
    return total_wins, total_losses, total


with open(f'../user_files/{username}_stat.txt', 'r') as f:
    stats = [UserStat(*line.rstrip().split('\t')) for line in f.readlines()]

totals = get_totals(stats)
list_2p = [row for row in stats if int(row.count) == 2]
list_3p = [row for row in stats if int(row.count) != 2]
totals2p = get_totals(list_2p)
totals3p = get_totals(list_3p)
print(totals, totals2p, totals3p)

totals_p = get_double_percentage(totals)
totals2p_p = get_double_percentage(totals2p)
totals3p_p = get_double_percentage(totals3p)
print(totals_p, totals2p_p, totals3p_p)

df1 = pd.DataFrame({'Count': ['wins', 'losses', 'total'], '2p': totals2p, '3p+': totals3p, 'Total': totals})
df2 = pd.DataFrame({'%': ['wins', 'losses'], '2p': totals2p_p, '3p+': totals3p_p, 'Total': totals_p})
writer = pd.ExcelWriter(f'../user_files/{username}_stats.xlsx', engine='xlsxwriter')
df1.to_excel(writer, sheet_name='stats', index=False, startrow=0)
df2.to_excel(writer, sheet_name='stats', index=False, startrow=5)
workbook = writer.book
worksheet = writer.sheets['stats']
border_fmt = workbook.add_format({'border': 1})
worksheet.conditional_format(xlsxwriter.utility.xl_range(0, 0, len(df1), len(df1.columns) - 1), {'type': 'no_errors', 'format': border_fmt})
worksheet.conditional_format(xlsxwriter.utility.xl_range(5, 0, 5 + len(df2), len(df2.columns) - 1), {'type': 'no_errors', 'format': border_fmt})
writer.save()
# workbook.add_format({
#         'bold': True,
#         'border': 1,
#         'text_wrap': True,
#         'align': 'center'
#     })
#
# wb = load_workbook(f'../user_files/{username}_stats.xlsx')
# print(wb.sheetnames)
